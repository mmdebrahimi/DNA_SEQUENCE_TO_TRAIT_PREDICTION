"""Score the FROZEN decoder on the 234-isolate Sci Rep 2023 cohort (2nd external cohort).

Source: PMC9829913 (Sci Rep 2023, BioProject PRJNA854358), open CC-BY supplement
MOESM1_ESM.xlsx (committed-out at data/raw/sci234/, gitignored). Everything is keyed by
study `Key`: "Supplementary data 1" = per-isolate GENOTYPE (QRDR POINT mutations gyrA_p*/
parC_p*/parE_p* + all acquired bla/aac/ant/aph/sul/dfr/... alleles as a wide presence table),
"Supplementary data 2" = per-isolate measured MIC (broth microdilution) incl CIPROF / GENTAM / CEFOTA.

Like Oxford (scripts/oxford_score.py): a pure JOIN of the cohort's OWN genotype + measured MIC,
scored by the frozen call_resistance RULE. NO download / assembly / Docker — the supplement ships
the genotype. Independence: independent genotype caller (ResFinder/PointFinder-style) + independent
measured MIC; MORE independent than NCBI-PD provdisjoint. Leakage DISJOINT by construction:
PRJNA854358 deposits 0 NCBI assemblies (reads-only); the decoder's tuning set is 100% GCA/GCF.

THREE drugs scored (all from ONE synthesized per-isolate main.tsv):
  - ciprofloxacin: QRDR-count rule (gyrA/parC/parE POINT mutations) — needs no Subclass lookup.
  - gentamicin:    GENTAMICIN-subclass refinement — acquired aac/ant/arm alleles resolved to their
                   AMRFinder Class/Subclass via the deployed fam.tsv (scripts/fam_subclass_resolver).
  - ceftriaxone:   CEPHALOSPORIN/CARBAPENEM extended-spectrum refinement — bla alleles resolved the
                   same way. CEFOTA (cefotaxime) MIC is the 3GC proxy (CLSI E. coli breakpoints
                   identical to ceftriaxone: S<=1 / R>=4).

FIDELITY CAVEAT (documented scope-limit): the synthesized main.tsv reproduces fam.tsv FAMILY-level
Class/Subclass curation, a faithful-but-imperfect proxy for real AMRFinder v4.2.7 node-level
assignment. CTX-M/CMY/DHA/ACT (the E. coli 3GC-R drivers) + NDM/KPC resolve correctly; rare
ESBL/carbapenemase variants whose fam.tsv family node is generic BETA-LACTAM (TEM-/SHV-ESBLs,
blaOXA-48) resolve to BETA-LACTAM and are NOT counted for cef -> a CONSERVATIVE under-call bounded
to rare variants. Any cef false-negative is surfaced in the artifact `discordances` for audit.
"""
from __future__ import annotations

import json
import re
import sys
import tempfile
from collections import Counter
from datetime import date as _date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dna_decode.data.mic_tiers import breakpoints_for
from dna_decode.eval.amr_rules import call_resistance
from scripts.fam_subclass_resolver import deployed_resolver
from scripts.independent_cohort_validate import _conf

SUPPL = Path("data/raw/sci234/sci234_supplement_MOESM1.xlsx")
REGISTRY_ORGANISM = "Escherichia_coli_Shigella"
_AMR_HDR = "Element symbol\tMethod\tClass\tSubclass\tElement name\t% Identity to reference"

# Genotype metadata columns in "Supplementary data 1" that are NOT resistance determinants.
_META_COLS = {"Key", "ESBL", "Collection date", "MLST PubMLST (Achtman) ST", "H_antigens", "O_antigens"}
# Drug -> (MIC column in Supplementary data 2, decoder drug for call_resistance, MIC-source note).
_DRUGS = {
    "ciprofloxacin": ("CIPROF", "ciprofloxacin", "CIPROF MIC"),
    "gentamicin":    ("GENTAM", "gentamicin", "GENTAM MIC"),
    "ceftriaxone":   ("CEFOTA", "ceftriaxone", "CEFOTA (cefotaxime) MIC — 3GC proxy; CLSI E.coli S<=1/R>=4 == ceftriaxone"),
}
# QRDR point-mutation column pattern (gyrA_pS83L style) -> AMRFinder symbol gyrA_S83L.
_QRDR_RE = re.compile(r"(gyrA|gyrB|parC|parE)_p")


def _present(v) -> bool:
    return v not in (None, 0, "0", "", "-", "NA", "NaN", False)


def _mic(cell) -> float | None:
    s = str(cell)
    for op in ("<=", ">=", ">", "<", "="):
        s = s.replace(op, "")
    s = s.strip()
    try:
        return float(s)
    except ValueError:
        return None


def load_cohort(suppl: Path = SUPPL):
    """Return (geno_qrdr, geno_acquired, mic_by_drug). geno_* keyed by study Key."""
    import openpyxl
    wb = openpyxl.load_workbook(suppl, read_only=True)
    # --- Supplementary data 1: genotype ---
    r1 = wb["Supplementary data 1"].iter_rows(values_only=True)
    h1 = list(next(r1))
    cols = [str(c) for c in h1]
    ki = cols.index("Key")
    qrdr_idx = [i for i, c in enumerate(cols) if _QRDR_RE.match(c)]
    # acquired-gene columns: every non-metadata, non-QRDR, non-point-mutation column (point muts carry
    # '_p<AA>' / '_d<nt>' patterns: rpoB_p, folP_p, pmrA_p, 16S_*_d, 23S_d, ampC_promoter_*; those are
    # NOT acquired genes and are not gent/cef determinants -> excluded by the '_p'/'_d'/promoter filter).
    def _is_point(c: str) -> bool:
        return ("_p" in c and re.search(r"_p[A-Z]\d", c) is not None) or "_d" in c or "promoter" in c
    acq_idx = [i for i, c in enumerate(cols)
               if c not in _META_COLS and i not in qrdr_idx and not _is_point(c)]
    geno_qrdr, geno_acq = {}, {}
    for r in r1:
        k = str(r[ki]).strip()
        geno_qrdr[k] = [cols[i].replace("_p", "_", 1) for i in qrdr_idx if i < len(r) and _present(r[i])]
        geno_acq[k] = [cols[i] for i in acq_idx if i < len(r) and _present(r[i])]
    # --- Supplementary data 2: measured MIC ---
    r2 = wb["Supplementary data 2"].iter_rows(values_only=True)
    h2 = list(next(r2))
    cols2 = [str(c) for c in h2]
    k2 = cols2.index("Key")
    mic_rows = [(str(r[k2]).strip(), r) for r in r2]
    mic_by_drug = {}
    for drug, (col, _dd, _note) in _DRUGS.items():
        ci = cols2.index(col) if col in cols2 else None
        mic_by_drug[drug] = {k: _mic(r[ci]) for k, r in mic_rows
                             if ci is not None and ci < len(r)} if ci is not None else {}
    return geno_qrdr, geno_acq, mic_by_drug


def binary_labels(mics: dict[str, float | None], drug: str) -> dict[str, str]:
    """R if MIC >= CLSI-R, S if MIC <= CLSI-S, intermediate/None excluded."""
    bp = breakpoints_for(drug)
    out = {}
    for k, m in mics.items():
        if m is None:
            continue
        if m >= bp["clsi_r"]:
            out[k] = "R"
        elif m <= bp["clsi_s"]:
            out[k] = "S"
    return out


def synthesize_main_tsv(qrdr: list[str], acquired: list[str], path: Path) -> dict:
    """Write a synthesized AMRFinder main.tsv for one isolate. Returns resolver stats.

    QRDR POINT rows -> Method POINTX, Class/Subclass QUINOLONE (the cipro counter keys on POINT + symbol).
    Acquired gene rows -> Method EXACTX, Class/Subclass from the deployed fam.tsv resolver (unresolved
    genes are dropped — they cannot be assigned a curated Subclass, so the conservative call is no-count)."""
    R = deployed_resolver()
    lines = [f"{s}\tPOINTX\tQUINOLONE\tQUINOLONE\t{s}\t100" for s in qrdr]
    unresolved = []
    for g in acquired:
        cls, sub, kind = R.resolve(g)
        if cls is None:
            unresolved.append(g)
            continue
        lines.append(f"{g}\tEXACTX\t{cls}\t{sub}\t{g}\t100")
    path.write_text(_AMR_HDR + "\n" + "\n".join(lines) + "\n", encoding="utf-8")
    return {"n_qrdr": len(qrdr), "n_acquired": len(acquired), "n_unresolved": len(unresolved),
            "unresolved": unresolved}


def score(geno_qrdr, geno_acq, mic_by_drug) -> dict:
    tmp = Path(tempfile.mkdtemp(prefix="sci234_"))
    # one synthesized main.tsv per isolate (shared across all drugs)
    main_paths, resolver_stats = {}, {}
    all_keys = set(geno_qrdr) | set(geno_acq)
    for k in all_keys:
        p = tmp / f"{k.replace('/', '_')}.tsv"
        resolver_stats[k] = synthesize_main_tsv(geno_qrdr.get(k, []), geno_acq.get(k, []), p)
        main_paths[k] = p

    results, discordances = {}, {}
    for drug, (col, dd, note) in _DRUGS.items():
        labels = binary_labels(mic_by_drug[drug], drug)
        pairs, disc = [], []
        for k, rs in labels.items():
            if k not in main_paths:
                continue
            res = call_resistance(main_paths[k], dd, organism=REGISTRY_ORGANISM)
            pred = res["prediction"]
            y = 1 if rs == "R" else 0
            pairs.append((pred, y))
            if (pred == "R") != bool(y) and pred != "ABSTAIN":
                disc.append({"key": k, "label": rs, "pred": pred, "n_det": res.get("n_determinants"),
                             "determinants": [d.get("symbol") for d in res.get("determinants", [])]})
        conf = _conf(pairs)
        conf["mic_source"] = note
        nR, nS = conf["tp"] + conf["fn"], conf["tn"] + conf["fp"]
        if nR >= 5 and nS >= 5:
            conf["powering"], conf["powering_reason"] = "SCORED", f"{nR}R/{nS}S both classes >=5"
        else:
            short = "no resistant isolates" if nR == 0 else ("no susceptible isolates" if nS == 0 else "minority class <5")
            conf["powering"] = "UNDERPOWERED"
            conf["powering_reason"] = f"{nR}R/{nS}S ({short}); sensitivity unmeasurable" if nR < 5 else f"{nR}R/{nS}S ({short})"
        results[drug] = conf
        discordances[drug] = disc
    # resolver coverage summary (all isolates pooled)
    unresolved_pool = Counter()
    for st in resolver_stats.values():
        unresolved_pool.update(st["unresolved"])
    return results, discordances, unresolved_pool


def main() -> int:
    geno_qrdr, geno_acq, mic_by_drug = load_cohort()
    n_iso = len(set(geno_qrdr) | set(geno_acq))
    print(f"loaded {n_iso} isolates; MIC keys per drug: "
          + ", ".join(f"{d}={len(m)}" for d, m in mic_by_drug.items()))
    results, discordances, unresolved_pool = score(geno_qrdr, geno_acq, mic_by_drug)
    for drug, c in results.items():
        print(f"\nSCI234 {drug}: [{c['powering']}] n={c['n_scored']} acc={c['acc']} sens={c['sens']} spec={c['spec']} "
              f"(R{c['tp']+c['fn']}/S{c['tn']+c['fp']}; abstain {c['abstain']}) — {c['powering_reason']}")
        d = discordances[drug]
        if d:
            fn = [x for x in d if x["label"] == "R"]
            fp = [x for x in d if x["label"] == "S"]
            print(f"  discordances: {len(fn)} FN (true-R called S), {len(fp)} FP (true-S called R)")
            for x in (fn + fp)[:6]:
                print(f"    {x['key']} label={x['label']} pred={x['pred']} n_det={x['n_det']} dets={x['determinants'][:6]}")
    if unresolved_pool:
        print(f"\nunresolved acquired genes (dropped; not counted): "
              f"{dict(unresolved_pool.most_common(12))}")

    art = {
        "_schema": "external-validation-v1", "date": _date.today().isoformat(),
        "cohort": "sci234_lanellaecoli_PRJNA854358", "organism": REGISTRY_ORGANISM,
        "drugs": {drug: {k: v for k, v in c.items()} for drug, c in results.items()},
        "discordances": discordances,
        "unresolved_acquired_genes": dict(unresolved_pool.most_common()),
        "evidence_tier": "external_clinical",
        "independence_tier": ("Sci Rep 2023 234-isolate E. coli cohort (Spain/Europe): own measured MIC "
                              "(broth microdilution) + own ResFinder/PointFinder-style genotype, joined by study Key. "
                              "Independent genotype caller; the call_resistance RULE is validated. cipro = QRDR-count "
                              "(no Subclass lookup); gent+cef use the deployed fam.tsv per-gene Subclass resolver "
                              "(scripts/fam_subclass_resolver.py) to reproduce AMRFinder v4.2.7 Class/Subclass."),
        "fidelity_caveat": ("gent+cef synthesize a main.tsv reproducing fam.tsv FAMILY-level Class/Subclass "
                            "(faithful-but-imperfect proxy for real AMRFinder node-level assignment). CTX-M/CMY/"
                            "DHA/ACT (E. coli 3GC drivers) + NDM/KPC resolve correctly; rare ESBL/carbapenemase "
                            "variants on a generic BETA-LACTAM family node (TEM/SHV-ESBL, blaOXA-48) resolve to "
                            "BETA-LACTAM and are NOT counted for cef -> conservative under-call, surfaced in "
                            "discordances. gent resolution is clean (aac(3)*/armA/ant(2'')-Ia = GENTAMICIN)."),
        "leakage_status": "DISJOINT_PROJECT_LEVEL",
        "leakage_evidence": ("PRJNA854358 deposits 0 NCBI assemblies (reads-only); decoder tuning set is 100% "
                             "GCA/GCF -> a 234-cohort isolate cannot be in it. Disjoint by accession-type construction."),
        "source": "PMC9829913 / Sci Rep 2023 MOESM1_ESM.xlsx (open CC-BY)",
        "note": "Direct supplement score (no assembly/Docker) — the genotype supplement ships QRDR mutations + acquired alleles.",
    }
    out = Path(f"wiki/external_validation_sci234_{_date.today().isoformat()}.json")
    out.write_text(json.dumps(art, indent=2), encoding="utf-8")
    print(f"\nartifact -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
